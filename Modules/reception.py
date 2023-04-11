import os
import sys
from datetime import date, time, datetime

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook
from sqlalchemy import or_, and_, exists
from starlette.responses import FileResponse

sys.path.append("..")

from fastapi import Depends, APIRouter, Request, Response, HTTPException, UploadFile, File
from Modules import auth
from sqlalchemy.orm import Session
from models import model
from schemas import schema

router = APIRouter(
    prefix="/reception",
    tags=["reception"],
    responses={401: {"user": "Not Authorised"}},
)


@router.post("/assign-persons-to-cultivator")
async def get_guests(obj: schema.assign_cultivator_to_persons, db: Session = Depends(auth.get_db)):
    for item in obj.persons:
        create_person_model = db.get(model.Person, item)
        create_person_model.cultivator_id = obj.cultivator.id
        db.add(create_person_model)
        db.commit()
    return {"msg": f"Persons Assigned to CULTIVATOR {obj.cultivator.name}"}


@router.get("/getcultivators")
async def get_cultivators(request: Request, db: Session = Depends(auth.get_db)):
    # user = await auth.get_current_user(request)
    # if user is None:
    #     raise HTTPException(status_code=401, detail="Sorry you are Unauthorized !")
    tempCultivators = db.query(model.User).filter(model.User.role_id == 2).offset(0).limit(
        50).all()
    cultivators: list = []
    for cultivator in tempCultivators:
        results = {"id": cultivator.person.id, "name": cultivator.person.name}
        cultivators.append(results)
    return {"cultivators": cultivators}


@router.get("/getguests")
async def get_guests(currentPage: int, pageSize: int, db: Session = Depends(auth.get_db)):
    offset = pageSize * (currentPage - 1)
    totalGuests = db.query(model.Person).filter(model.Person.cultivator_id == None).count()
    tempGuests = db.query(model.Person).filter(model.Person.cultivator_id == None).order_by(
        model.Person.id.desc()).offset(offset).limit(
        pageSize).all()
    guests: list = []
    for guest in tempGuests:
        results = {"id": guest.id, "name": guest.name, "phone_no": guest.phone_no, "email": guest.email,
                   "gender": guest.gender}
        guests.append(results)
    return {"guests": guests, "totalGuests": totalGuests}


@router.post("/savefromexcel")
async def save_from_excel(file: UploadFile = File(...), db: Session = Depends(auth.get_db)):
    # path = "C:\\Users\\dcdrns\\Desktop\\Books.xlsx" #syntax for path in python
    # filename, file_extension = os.path.splitext(file.filename) #split the file name and extension
    # file.filename = f"tempXLfile{file_extension}"  # rename the file

    file.filename = f"tempXLfile.xlsx"  # rename the file
    contents = await file.read()
    with open(f"{file.filename}", "wb") as f:
        f.write(contents)
    try:
        wb_obj = openpyxl.load_workbook("tempXLfile.xlsx")
    except:
        return {"msg": "Not a valid .xlsx file !"}

    check_fields = ["name", "first_name", "last_name", "phone_no", "email", "city", "gender", "dob"]

    sheet_obj = wb_obj.active  # This is set to 0 by default. Unless you modify its value, you will always get the
    # first worksheet by using this method.

    for i in range(2, sheet_obj.max_row + 1):  # row
        if sheet_obj.cell(row=i, column=4).data_type != 'n' \
                or sheet_obj.cell(row=i, column=4).value is None \
                or sheet_obj.cell(row=i,
                                  column=1).value is None:  # name cannot be null phone_no cant' be null or non numeric
            return {
                "msg": f"Emty or Invalid Name and phone no. in Row: {i}. \n max_row: {sheet_obj.max_row} \n max_column: {sheet_obj.max_column}"}

    for i in range(1, sheet_obj.max_row + 1):  # row

        if i == 1:
            for j in range(1, sheet_obj.max_column + 1):  # itrate each column
                # cell_obj = sheet_obj['A2']
                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value != check_fields[j - 1] or sheet_obj.max_column > len(check_fields):
                    return {"msg": ".xlsx file sample not correct."}
        else:
            person = model.Person()
            person.name = sheet_obj.cell(row=i, column=1).value
            person.first_name = sheet_obj.cell(row=i, column=2).value
            person.last_name = sheet_obj.cell(row=i, column=3).value
            person.phone_no = sheet_obj.cell(row=i, column=4).value
            person.email = sheet_obj.cell(row=i, column=5).value
            person.city = sheet_obj.cell(row=i, column=6).value
            person.gender = sheet_obj.cell(row=i, column=7).value
            # person.dob = '1997-11-11 13:23:44'
            if sheet_obj.cell(row=i, column=8).is_date:
                person.dob = sheet_obj.cell(row=i, column=8).value

            db_person = db.query(model.Person).filter(
                model.Person.name == person.name).filter(model.Person.phone_no == person.phone_no).first()
            if db_person:
                return {
                    "msg": f"Person with name'{person.name}' and phone '{person.phone_no}', already Exist in Row: {i}."}

            db.add(person)
            db.commit()
            db.refresh(person)
    wb_obj.close()  # not necessary
    return {"msg": "Excel sheet uploaded !"}


@router.get("/search")
async def search(currentPage: int = 1, pageSize: int = 10, search_input: str = '', db: Session = Depends(auth.get_db)):
    searchData = f'%{search_input}%'
    if searchData == '%%':
        # return {"msg": "Empty search value."}
        raise HTTPException(status_code=404, detail="Empty Search value!")
    offset = pageSize * (currentPage - 1)

    if search_input.isnumeric():
        persons = db.query(model.Person).filter(and_(model.Person.cultivator_id == None,model.Person.phone_no.ilike(
                    searchData))).offset(offset).limit(pageSize).all()
                  # ilike gurantees case insensitive
        totalGuests = db.query(model.Person).filter(and_(model.Person.cultivator_id == None,model.Person.phone_no.ilike(
                    searchData))).count()
    else:
        persons = db.query(model.Person).filter(and_(model.Person.cultivator_id == None, model.Person.name.ilike(
            searchData))).offset(offset).limit(
            pageSize).all()
        # ilike gurantees case insensitive
        totalGuests = db.query(model.Person).filter(and_(model.Person.cultivator_id == None, model.Person.name.ilike(
            searchData))).count()

    return {"persons": persons, "totalGuests": totalGuests}


@router.post("/createorientation")
async def create_orientation(orientation: schema.create_orientation, db: Session = Depends(auth.get_db)):
    orientation_temp = model.Orientation(**orientation.dict())
    db.add(orientation_temp)
    db.commit()
    db.refresh(orientation_temp)
    print(orientation)
    return {"msg": orientation}


@router.get("/getorientations")
async def get_orientations(db: Session = Depends(auth.get_db)):
    orientations = db.query(model.Orientation).with_entities(model.Orientation.id,
                                                             model.Orientation.orientation_name,
                                                             model.Orientation.orientation_end_date_time,
                                                             model.Orientation.orientation_start_date_time,
                                                             model.Orientation.venue,
                                                             model.Person.name.label("cultivator_name")
                                                             # select person.name as cultivator_name
                                                             ).filter(
        model.Orientation.cultivator_id == model.Person.id).all()
    return {"orientations": orientations}


@router.post("/createvisits")
async def create_visits(file: UploadFile = File(...), db: Session = Depends(auth.get_db)):
    file.filename = f"tempXLfile1.xlsx"  # rename the file
    contents = await file.read()
    with open(f"{file.filename}", "wb") as f:
        f.write(contents)
    try:
        wb_obj = openpyxl.load_workbook("tempXLfile1.xlsx")
    except:
        return {"msg": "Not a valid .xlsx file !"}

    check_fields = ["name", "phone_no", "check_in_date_time", "check_out_date_time", "accommodation"]

    sheet_obj = wb_obj.active  # This is set to 0 by default. Unless you modify its value, you will always get the
    # first worksheet by using this method.

    for i in range(2, sheet_obj.max_row + 1):  # row
        if sheet_obj.cell(row=i, column=2).data_type != 'n' \
                or sheet_obj.cell(row=i, column=1).value is None \
                or sheet_obj.cell(row=i, column=2).value is None \
                or sheet_obj.cell(row=i, column=3).value is None:
            return {
                "msg": f"Emty or Invalid Fields in Row: {i}. \n max_row: {sheet_obj.max_row} \n max_column: {sheet_obj.max_column}"}
    for i in range(1, sheet_obj.max_row + 1):  # row
        if i == 1:
            for j in range(1, sheet_obj.max_column + 1):  # itrate each column
                # cell_obj = sheet_obj['A2']
                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value != check_fields[j - 1] or sheet_obj.max_column > len(check_fields):
                    return {"msg": ".xlsx file sample not correct."}
        else:
            name = sheet_obj.cell(row=i, column=1).value
            phone = sheet_obj.cell(row=i, column=2).value
            start_time = sheet_obj.cell(row=i, column=3).value
            end_time = sheet_obj.cell(row=i, column=4).value
            db_person = db.query(model.Person).filter(
                and_(model.Person.name == name, model.Person.phone_no == phone)).first()
            if not db_person:  # if person does not exist
                db_person = model.Person(
                    name=name,
                    phone_no=phone
                )
                db.add(db_person)
                db.commit()
                db.refresh(db_person)

            db_visit = model.Visit(
                person_id=db_person.id,
                check_in_date_time=start_time,
                check_out_date_time=end_time
            )
            db.add(db_visit)
            db.commit()
            db.refresh(db_visit)
    wb_obj.close()  # not necessary
    return {"msg": "Visits Created", "filename": file.filename}


@router.get("/getvisits")
async def get_visits(db: Session = Depends(auth.get_db)):
    visits = db.query(model.Visit).with_entities(
        model.Visit.id,
        model.Visit.check_in_date_time,
        model.Person.name.label("visitor_name"),
        model.Person.phone_no
    ).filter(
        and_(model.Visit.person_id == model.Person.id, model.Visit.orientation_assigned == False)).all()
    return {"visits": visits}


@router.post("/addparticipants")
async def add_participants(obj: schema.addparticipants, db: Session = Depends(auth.get_db)):
    for visit in obj.visitors:
        visit_update_orientation_assigned = db.get(model.Visit, visit.get("visit_id"))
        visit_update_orientation_assigned.orientation_assigned = True

        participant = model.OrientationParticipants(
            orientation_id=obj.orientation,
            visit_id=visit.get("visit_id"),
            visitor_name=visit.get("visitor_name")
        )
        db.add(participant)
        db.commit()
        db.refresh(participant)
        db.add(visit_update_orientation_assigned)
        db.commit()
        db.refresh(visit_update_orientation_assigned)

    return {"msg": "participants added"}


@router.get("/getparticipants")
async def get_participants(orientation_id: int, db: Session = Depends(auth.get_db)):
    participants = db.query(model.OrientationParticipants).filter(
        model.OrientationParticipants.orientation_id == orientation_id).all()
    return {"participants": participants}


@router.post("/createexcelsheet")
async def create_excel_sheet(exportxcel: schema.exportexcel, db: Session = Depends(auth.get_db)):
    workbook = Workbook()
    sheet = workbook.active
    field_names = ["id", "name", "phone_no", "email", "gender", "dob"]
    for i in range(1, len(field_names) + 1):
        sheet.cell(row=1, column=i).value = field_names[i - 1]
        sheet.cell(row=1, column=i).font = Font(bold=True)
        sheet.cell(row=1, column=i).alignment = Alignment(horizontal='center')

    for j in range(1, len(exportxcel.persons) + 1):
        db_person = db.query(model.Person).with_entities(model.Person.id, model.Person.name, model.Person.phone_no,
                                                         model.Person.email, model.Person.gender,
                                                         model.Person.dob).filter(
            model.Person.id == exportxcel.persons[j - 1]).first()
        print(db_person[1])
        for i in range(1, len(db_person) + 1):
            sheet.cell(row=j + 1, column=i).value = db_person[i - 1]
    workbook.save("export.xlsx")
    return FileResponse("export.xlsx")


