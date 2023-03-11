import os
import sys

import openpyxl

sys.path.append("..")

from Modules import auth
from fastapi import Depends, APIRouter, Request, Response, HTTPException, UploadFile, File
from fastapi.templating import Jinja2Templates
from models import model
from Common import services
from sqlalchemy.orm import Session
from schemas import schema

templates = Jinja2Templates(directory="static")
router = APIRouter(
    prefix="/test",
    tags=["test"],
    responses={401: {"user": "Not Authorised"}},
)


@router.post("/exceltest")
async def excel_test(file: UploadFile = File(...), db: Session = Depends(auth.get_db)):
    # path = "C:\\Users\\dcdrns\\Desktop\\Books.xlsx" #syntax for path in python
    # filename, file_extension = os.path.splitext(file.filename) #split the file name and extension
    # file.filename = f"tempXLfile{file_extension}"  # rename the file

    file.filename = f"tempXLfile.xlsx"  # rename the file
    contents = await file.read()
    with open(f"{file.filename}", "wb") as f:
        f.write(contents)
    try:
        wb_obj = openpyxl.load_workbook("tempXLfile.xlsx")
    except:
        return {"msg": "Not a valid .xlsx file !"}

    check_fields = ["name", "first_name", "last_name", "phone_no", "email", "city", "gender", "dob"]

    sheet_obj = wb_obj.active  # This is set to 0 by default. Unless you modify its value, you will always get the
    # first worksheet by using this method.

    for i in range(2, sheet_obj.max_row + 1):  # row
        if sheet_obj.cell(row=i, column=4).data_type != 'n' \
                or sheet_obj.cell(row=i, column=4).value is None \
                or sheet_obj.cell(row=i,
                                  column=1).value is None:  # name cannot be null phone_no cant' be null or non numeric
            return {"msg": f"Emty or Invalid Name and phone no. in Row: {i}. \n max_row: {sheet_obj.max_row} \n max_column: {sheet_obj.max_column}"}

    for i in range(1, sheet_obj.max_row + 1):  # row

        if i == 1:
            for j in range(1, sheet_obj.max_column + 1):  # itrate each column
                # cell_obj = sheet_obj['A2']
                cell_obj = sheet_obj.cell(row=i, column=j)
                if cell_obj.value != check_fields[j - 1] or sheet_obj.max_column > len(check_fields):
                    return {"msg": ".xlsx file sample not correct."}
        else:
            person = model.Person()
            person.name = sheet_obj.cell(row=i, column=1).value
            person.first_name = sheet_obj.cell(row=i, column=2).value
            person.last_name = sheet_obj.cell(row=i, column=3).value
            person.phone_no = sheet_obj.cell(row=i, column=4).value
            person.email = sheet_obj.cell(row=i, column=5).value
            person.city = sheet_obj.cell(row=i, column=6).value
            person.gender = sheet_obj.cell(row=i, column=7).value
            # person.dob = '1997-11-11 13:23:44'
            if sheet_obj.cell(row=i, column=8).is_date:
                person.dob = sheet_obj.cell(row=i, column=8).value

            db_person = db.query(model.Person).filter(
                model.Person.name == person.name).filter(model.Person.phone_no == person.phone_no).first()
            if db_person:
                return {"msg": f"Person with same name and phone no. Already Exist. Row: {i}."}

            db.add(person)
            db.commit()
            db.refresh(person)
    wb_obj.close()  # not necessary
    return {"msg": "Excel sheet uploaded !"}


# --------------------------pydentic validator test-------------------------------------------------------------
@router.post("/testValidator")
async def serve_my_app(test: schema.test, request: Request):
    return {"test": test}


@router.post("/testalchemy")
async def test_alchemy(test: schema.test, db: Session = Depends(services.get_database_session)):
    # query = select([model.User.username]).where(model.User.id.in_([15]))
    # users = auth.engine.execute(query).all()
    # tempUsers = db.query(model.User.username, model.User.id).where(model.User.id.in_([19])).all()
    tempUsers = db.query(model.Role.id).all()
    users: list = [{"hi": "hello", "arrey": [1, 2, 5, 6]}]
    for user in tempUsers:
        result = user.id
        users.append(result)
    return {"users": users, "object": test}


# //-----dependency test ----------------------------------------
dummy = "true"


async def test1():
    if dummy != 'tgrue':
        raise HTTPException(status_code=401, detail="Sorry you are Unauthorized !")
    return "true"


@router.get("/dependencytest")
async def test_alchemy(temp: str = Depends(test1)):
    return temp


# //-----Relationship tables test----------------------------------------
# @router.get("/createindividual")
# async def create_individual(name: str, phone_no: str, db: Session = Depends(auth.get_db)):
#     individual = model.Individual(name=name, phone_no=phone_no)
#     db.add(individual)
#     db.commit()
#     return "Individual created"
#

@router.get("/createrelationship")
async def create_relationship(relation_name: str, person_primary_id: int, person_secondary_id: int,
                              reverse_relationship_name: str, db: Session = Depends(auth.get_db)):
    check_relation = db.query(model.PersonRelationships).filter(
        model.PersonRelationships.person_primary_id == person_primary_id).filter(
        model.PersonRelationships.person_secondary_id == person_secondary_id).first()
    if check_relation:
        raise HTTPException(status_code=404, detail=f"Relation Between these two persons already Exist")

    create_relation = model.PersonRelationships(relationship_name=relation_name, person_primary_id=person_primary_id,
                                                person_secondary_id=person_secondary_id,
                                                reverse_relationship_name=reverse_relationship_name)
    db.add(create_relation)
    db.commit()
    return "Relation created"


# @router.get("/getfamilypersons") async def get_family_persons(Id: int, db: Session = Depends(auth.get_db)):
# family_relations = db.query(model.PersonRelationships).filter(model.PersonRelationships.person_primary_id ==
# Id).offset( 0).limit( 50).all() family_members: list = [] for item in family_relations: person_primary = db.get(
# model.Person, item.person_primary_id) person_secondary = db.get(model.Person, item.person_secondary_id) result = {
# "self": person_primary.name, "person_name": person_secondary.name, "relation": item.relationship_name}
# family_members.append(result) return family_members

@router.get("/getfamilypersons")
async def get_family_persons(Id: int, db: Session = Depends(auth.get_db)):
    family_relations_fwd = db.query(model.PersonRelationships).filter(
        model.PersonRelationships.person_primary_id == Id).all()
    family_relations_bkwrd = db.query(model.PersonRelationships).filter(
        model.PersonRelationships.person_secondary_id == Id).all()

    for item_fwd in family_relations_fwd:
        for item_bkwrd in family_relations_bkwrd:
            if item_fwd.person_secondary_id == item_bkwrd.person_primary_id:
                family_relations_bkwrd.remove(item_bkwrd)
    family_relation_arrey = family_relations_fwd + family_relations_bkwrd

    family_relations: list = []
    for item in family_relation_arrey:
        person_primary = db.get(model.Person, item.person_primary_id)
        person_secondary = db.get(model.Person, item.person_secondary_id)
        if item.person_primary_id == Id:
            family_relations.append({"primary_name": person_primary.name, "secondary_name": person_secondary.name,
                                     "relation_name": item.relationship_name,
                                     "msg": f"{person_primary.name}'s {item.relationship_name} is {person_secondary.name}.",
                                     "primary_id": item.person_primary_id, "secondary_id": item.person_secondary_id})
        else:
            family_relations.append({"primary_name": person_secondary.name, "secondary_name": person_primary.name,
                                     "relation_name": item.reverse_relationship_name,
                                     "msg": f"{person_secondary.name}'s {item.reverse_relationship_name} is {person_primary.name}. (predicted)",
                                     "primary_id": item.person_primary_id, "secondary_id": item.person_secondary_id})

    return {"family_relations": family_relations}
